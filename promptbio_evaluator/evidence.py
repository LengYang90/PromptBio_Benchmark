"""Evidence inventory and read-only content access for Deep Agent tools."""

from __future__ import annotations

import base64
import csv
import hashlib
import io
import json
import math
import mimetypes
import statistics
from dataclasses import dataclass
from difflib import SequenceMatcher
from pathlib import Path
from threading import RLock
from typing import Any, Literal

from .models import EvidenceRecord, TaskSpec
from .policy import AllowedFile, PolicyError, Stage, file_sha256, stage_allowed_files


TEXT_EXTENSIONS = {
    ".txt", ".csv", ".tsv", ".json", ".jsonl", ".yaml", ".yml", ".xml",
    ".html", ".htm", ".md", ".rst", ".log", ".out", ".err", ".sh", ".py",
    ".r", ".rmd", ".pl", ".awk", ".bed", ".gff", ".gff3", ".gtf", ".vcf",
    ".fasta", ".fa", ".fna", ".ffn", ".faa", ".frn", ".fastq", ".fq", ".sam",
    # STAR Chimeric.out.junction is a tab-delimited text report. It has no
    # required header, so read_text/search_text are safer than header-based
    # table sampling for this artifact type.
    ".junction",
}
TABLE_EXTENSIONS = {".csv", ".tsv", ".txt", ".xlsx", ".xlsm"}
IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".gif", ".webp", ".tif", ".tiff"}
BIO_EXTENSIONS = {".fa", ".fasta", ".fna", ".ffn", ".faa", ".frn", ".fq", ".fastq", ".vcf", ".bcf", ".bam", ".cram"}
PDB_EXTENSION = ".pdb"

THREE_TO_ONE = {
    "ALA": "A", "ARG": "R", "ASN": "N", "ASP": "D", "CYS": "C",
    "GLN": "Q", "GLU": "E", "GLY": "G", "HIS": "H", "ILE": "I",
    "LEU": "L", "LYS": "K", "MET": "M", "PHE": "F", "PRO": "P",
    "SER": "S", "THR": "T", "TRP": "W", "TYR": "Y", "VAL": "V",
    # Common modified amino acids that still identify the same polymer residue.
    "MSE": "M", "SEC": "U", "PYL": "O",
}


@dataclass(frozen=True)
class AccessLimits:
    """Bound the size of a single read while allowing explicit pagination."""

    max_text_characters: int = 1_000_000
    max_table_rows: int = 200
    max_records: int = 100
    max_image_bytes: int = 8_000_000


class EvidenceStore:
    """Stage-scoped, immutable source catalog plus an append-only evidence ledger."""

    def __init__(
        self,
        task_dir: Path,
        result_dir: Path,
        task: TaskSpec,
        stage: Stage,
        limits: AccessLimits,
    ) -> None:
        self.task_dir = Path(task_dir).resolve()
        self.result_dir = Path(result_dir).resolve()
        self.task = task
        self.stage = stage
        self.limits = limits
        allowed = stage_allowed_files(self.task_dir, self.result_dir, task, stage)
        self._files: dict[Path, AllowedFile] = {item.path: item for item in allowed}
        self._records: list[EvidenceRecord] = []
        self._partial_paths: set[Path] = set()
        self._hashes: dict[Path, str] = {}
        self._lock = RLock()
        self._next_evidence_number = 1
        self._evidence_prefix = "I" if stage == "initial_assessment" else "A"

    def _relative_path(self, path: Path) -> str:
        for root in (self.task_dir, self.result_dir):
            try:
                return str(path.relative_to(root.parent))
            except ValueError:
                continue
        return str(path)

    def _resolve_allowed(self, raw_path: str) -> tuple[Path, AllowedFile]:
        candidate = Path(raw_path)
        if not candidate.is_absolute():
            # Tool results list absolute paths, but resolving a relative path
            # against the task root makes manual CLI inspection predictable.
            candidate = self.task_dir / candidate
        try:
            resolved = candidate.resolve(strict=True)
        except FileNotFoundError as exc:
            raise PolicyError(f"Evidence file does not exist: {raw_path}") from exc
        entry = self._files.get(resolved)
        if entry is None:
            raise PolicyError(
                f"Path is outside the {self.stage} read-only allowlist: {raw_path}"
            )
        return resolved, entry

    def _hash(self, path: Path) -> str:
        with self._lock:
            if path not in self._hashes:
                self._hashes[path] = file_sha256(path)
            return self._hashes[path]

    def cite(
        self,
        path: Path,
        *,
        locator: str,
        extractor: str,
        status: Literal["complete", "partial"],
        notes: list[str] | None = None,
    ) -> EvidenceRecord:
        """Record a content inspection with a thread-safe, unique ID."""
        with self._lock:
            evidence_id = f"{self._evidence_prefix}-{self._next_evidence_number:04d}"
            self._next_evidence_number += 1
            if status == "partial":
                self._partial_paths.add(path)
            record = EvidenceRecord(
                evidence_id=evidence_id,
                path=self._relative_path(path),
                locator=locator,
                extractor=extractor,
                sha256=self._hash(path),
                status=status,
                notes=notes or [],
            )
            self._records.append(record)
            return record

    def allowed_file_manifest(self) -> list[dict[str, Any]]:
        """List navigation metadata, without creating a citation record."""
        entries: list[dict[str, Any]] = []
        for path, item in sorted(self._files.items()):
            entries.append(
                {
                    "path": str(path),
                    "role": item.role,
                    "size_bytes": path.stat().st_size,
                    "suffix": path.suffix.lower(),
                    "mime_type": mimetypes.guess_type(path.name)[0] or "application/octet-stream",
                    "sha256": self._hash(path),
                }
            )
        return entries

    def metadata(self, raw_path: str) -> dict[str, Any]:
        path, item = self._resolve_allowed(raw_path)
        return {
            "path": str(path),
            "role": item.role,
            "size_bytes": path.stat().st_size,
            "suffix": path.suffix.lower(),
            "mime_type": mimetypes.guess_type(path.name)[0] or "application/octet-stream",
            "sha256": self._hash(path),
        }

    def report_file_manifest(self) -> list[dict[str, Any]]:
        """Return each source once for report-level file provenance."""
        return [
            {
                "path": self._relative_path(path),
                "role": item.role,
                "stage": self.stage,
                "size_bytes": path.stat().st_size,
                "suffix": path.suffix.lower(),
                "mime_type": mimetypes.guess_type(path.name)[0] or "application/octet-stream",
                "sha256": self._hash(path),
            }
            for path, item in sorted(self._files.items())
        ]

    def read_text(self, raw_path: str, start_line: int, line_count: int) -> dict[str, Any]:
        if start_line < 1 or line_count < 1:
            raise PolicyError("start_line and line_count must be positive integers")
        path, _ = self._resolve_allowed(raw_path)
        if path.suffix.lower() not in TEXT_EXTENSIONS:
            raise PolicyError(
                f"{path.name} is not a supported plain-text file; choose a type-specific inspection tool"
            )

        selected: list[str] = []
        total_lines = 0
        total_characters = 0
        was_truncated = False
        last_line = start_line - 1
        with path.open("r", encoding="utf-8", errors="replace") as handle:
            for line_number, line in enumerate(handle, start=1):
                total_lines = line_number
                if line_number < start_line:
                    continue
                if len(selected) >= line_count:
                    was_truncated = True
                    continue
                if total_characters + len(line) > self.limits.max_text_characters:
                    was_truncated = True
                    break
                selected.append(line)
                total_characters += len(line)
                last_line = line_number

        locator = f"lines {start_line}-{last_line} of {total_lines}" if selected else f"lines {start_line}-{start_line} (empty range)"
        citation = self.cite(
            path,
            locator=locator,
            extractor="read_text",
            status="partial" if was_truncated else "complete",
            notes=["Output was limited; request the next line range for additional text."] if was_truncated else [],
        )
        return {
            "path": str(path),
            "content": "".join(selected),
            "start_line": start_line,
            "end_line": last_line,
            "total_lines_observed": total_lines,
            "truncated": was_truncated,
            "evidence_id": citation.evidence_id,
        }

    def search_text(
        self, raw_path: str, query: str, context_lines: int, max_matches: int
    ) -> dict[str, Any]:
        if not query:
            raise PolicyError("query must not be empty")
        if context_lines < 0 or max_matches < 1:
            raise PolicyError("context_lines must be non-negative and max_matches must be positive")
        path, _ = self._resolve_allowed(raw_path)
        if path.suffix.lower() not in TEXT_EXTENSIONS:
            raise PolicyError("search_text only supports plain-text evidence files")

        lines = path.read_text(encoding="utf-8", errors="replace").splitlines()
        matches: list[dict[str, Any]] = []
        for index, line in enumerate(lines):
            if query not in line:
                continue
            start = max(0, index - context_lines)
            end = min(len(lines), index + context_lines + 1)
            matches.append(
                {
                    "line": index + 1,
                    "context_start_line": start + 1,
                    "context_end_line": end,
                    "content": "\n".join(lines[start:end]),
                }
            )
            if len(matches) >= max_matches:
                break
        truncated = sum(1 for line in lines if query in line) > len(matches)
        citation = self.cite(
            path,
            locator=f"search query {query!r}, {len(matches)} match(es)",
            extractor="search_text",
            status="partial" if truncated else "complete",
            notes=["Additional matches were omitted."] if truncated else [],
        )
        return {
            "path": str(path),
            "query": query,
            "matches": matches,
            "truncated": truncated,
            "evidence_id": citation.evidence_id,
        }

    def inspect_json(self, raw_path: str, pointer: str) -> dict[str, Any]:
        path, _ = self._resolve_allowed(raw_path)
        try:
            value: Any = json.loads(path.read_text(encoding="utf-8"))
        except json.JSONDecodeError as exc:
            raise PolicyError(f"Invalid JSON file {path.name}: {exc}") from exc
        if pointer and pointer != "/":
            for part in pointer.lstrip("/").split("/"):
                part = part.replace("~1", "/").replace("~0", "~")
                if isinstance(value, list):
                    value = value[int(part)]
                elif isinstance(value, dict):
                    value = value[part]
                else:
                    raise PolicyError(f"JSON pointer cannot descend into {type(value).__name__}")
        rendered = json.dumps(value, ensure_ascii=False, indent=2)
        truncated = len(rendered) > self.limits.max_text_characters
        content = rendered[: self.limits.max_text_characters] if truncated else rendered
        citation = self.cite(
            path,
            locator=f"JSON pointer {pointer or '/'}",
            extractor="inspect_json",
            status="partial" if truncated else "complete",
            notes=["Selected JSON value was truncated."] if truncated else [],
        )
        return {"path": str(path), "pointer": pointer or "/", "content": content, "truncated": truncated, "evidence_id": citation.evidence_id}

    def inspect_table(self, raw_path: str, sample_rows: int) -> dict[str, Any]:
        if sample_rows < 1:
            raise PolicyError("sample_rows must be positive")
        path, _ = self._resolve_allowed(raw_path)
        suffix = path.suffix.lower()
        row_limit = min(sample_rows, self.limits.max_table_rows)
        if suffix in {".xlsx", ".xlsm"}:
            return self._inspect_excel(path, row_limit)
        if suffix not in {".csv", ".tsv", ".txt"}:
            raise PolicyError("inspect_table supports CSV, TSV, TXT, XLSX, and XLSM files")
        delimiter = "\t" if suffix == ".tsv" else ","
        with path.open("r", encoding="utf-8", errors="replace", newline="") as handle:
            reader = csv.reader(handle, delimiter=delimiter)
            rows = list(reader)
        header = rows[0] if rows else []
        sample = rows[1 : row_limit + 1]
        citation = self.cite(
            path,
            locator=f"header and first {len(sample)} data rows of {max(len(rows) - 1, 0)}",
            extractor="inspect_table",
            status="partial" if len(rows) - 1 > len(sample) else "complete",
            notes=["Use read_text for other row ranges."] if len(rows) - 1 > len(sample) else [],
        )
        return {
            "path": str(path),
            "format": suffix.lstrip("."),
            "header": header,
            "row_count": max(len(rows) - 1, 0),
            "sample_rows": sample,
            "truncated": len(rows) - 1 > len(sample),
            "evidence_id": citation.evidence_id,
        }

    def _inspect_excel(self, path: Path, row_limit: int) -> dict[str, Any]:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise PolicyError("Excel inspection requires openpyxl; install requirements.txt") from exc
        workbook = load_workbook(path, read_only=True, data_only=False)
        sheets: list[dict[str, Any]] = []
        for worksheet in workbook.worksheets:
            rows = []
            for row in worksheet.iter_rows(values_only=False):
                rows.append([cell.value for cell in row])
                if len(rows) >= row_limit + 1:
                    break
            sheets.append(
                {
                    "name": worksheet.title,
                    "max_row": worksheet.max_row,
                    "max_column": worksheet.max_column,
                    "header": rows[0] if rows else [],
                    "sample_rows": rows[1:],
                }
            )
        workbook.close()
        citation = self.cite(
            path,
            locator=f"all sheets; up to {row_limit} data rows per sheet",
            extractor="inspect_table",
            status="partial",
            notes=["Workbook inspection is sampled by row limit."],
        )
        return {"path": str(path), "format": "excel", "sheets": sheets, "truncated": True, "evidence_id": citation.evidence_id}

    def inspect_image(self, raw_path: str, include_preview: bool) -> list[dict[str, Any]]:
        path, _ = self._resolve_allowed(raw_path)
        if path.suffix.lower() not in IMAGE_EXTENSIONS:
            raise PolicyError("inspect_image only supports common raster image files")
        try:
            from PIL import Image
        except ImportError as exc:
            raise PolicyError("Image inspection requires Pillow; install requirements.txt") from exc

        with Image.open(path) as image:
            metadata = {"format": image.format, "width": image.width, "height": image.height, "mode": image.mode}
            preview: str | None = None
            if include_preview:
                image.thumbnail((1600, 1600))
                rendered = io.BytesIO()
                image.convert("RGB").save(rendered, format="PNG", optimize=True)
                payload = rendered.getvalue()
                if len(payload) <= self.limits.max_image_bytes:
                    preview = "data:image/png;base64," + base64.b64encode(payload).decode("ascii")
        citation = self.cite(
            path,
            locator="full image metadata" + (" and scaled preview" if preview else ""),
            extractor="inspect_image",
            status="complete",
            notes=["Preview omitted because it exceeded the configured byte limit."] if include_preview and not preview else [],
        )
        blocks: list[dict[str, Any]] = [{"type": "text", "text": json.dumps({**metadata, "path": str(path), "evidence_id": citation.evidence_id})}]
        if preview:
            blocks.append({"type": "image_url", "image_url": {"url": preview}})
        return blocks

    def inspect_pdf(self, raw_path: str, pages: list[int], include_preview: bool) -> list[dict[str, Any]]:
        path, _ = self._resolve_allowed(raw_path)
        if path.suffix.lower() != ".pdf":
            raise PolicyError("inspect_pdf only supports PDF files")
        try:
            import fitz  # PyMuPDF
        except ImportError as exc:
            raise PolicyError("PDF inspection requires PyMuPDF; install requirements.txt") from exc
        document = fitz.open(path)
        requested = pages or list(range(1, min(document.page_count, 5) + 1))
        output: list[dict[str, Any]] = []
        citations: list[str] = []
        for page_number in requested:
            if page_number < 1 or page_number > document.page_count:
                raise PolicyError(f"PDF page {page_number} is out of range 1-{document.page_count}")
            page = document.load_page(page_number - 1)
            text = page.get_text("text")
            text_truncated = len(text) > self.limits.max_text_characters
            citation = self.cite(
                path,
                locator=f"page {page_number} of {document.page_count}",
                extractor="inspect_pdf",
                status="partial" if text_truncated else "complete",
                notes=["Page text was truncated."] if text_truncated else [],
            )
            citations.append(citation.evidence_id)
            output.append({"type": "text", "text": json.dumps({"page": page_number, "page_count": document.page_count, "text": text[: self.limits.max_text_characters], "evidence_id": citation.evidence_id})})
            if include_preview:
                pixmap = page.get_pixmap(matrix=fitz.Matrix(1.5, 1.5), alpha=False)
                png = pixmap.tobytes("png")
                if len(png) <= self.limits.max_image_bytes:
                    output.append({"type": "image_url", "image_url": {"url": "data:image/png;base64," + base64.b64encode(png).decode("ascii")}})
        document.close()
        return output

    def inspect_bio(self, raw_path: str, record_limit: int) -> dict[str, Any]:
        if record_limit < 1:
            raise PolicyError("record_limit must be positive")
        path, _ = self._resolve_allowed(raw_path)
        suffix = path.suffix.lower()
        if suffix not in BIO_EXTENSIONS:
            raise PolicyError("inspect_bio supports FASTA/FASTQ, VCF/BCF, BAM, and CRAM files")
        record_limit = min(record_limit, self.limits.max_records)
        if suffix in {".fa", ".fasta", ".fna", ".ffn", ".faa", ".frn", ".fq", ".fastq"}:
            return self._inspect_sequence_file(path, record_limit)
        if suffix in {".vcf", ".bcf"}:
            return self._inspect_variant_file(path, record_limit)
        return self._inspect_alignment_file(path, record_limit)

    @staticmethod
    def _summary(values: list[float]) -> dict[str, float | int] | None:
        if not values:
            return None
        return {
            "count": len(values),
            "mean": round(statistics.fmean(values), 3),
            "median": round(statistics.median(values), 3),
            "minimum": round(min(values), 3),
            "maximum": round(max(values), 3),
        }

    @staticmethod
    def _sequence_preview(sequence: str, limit: int) -> str:
        if len(sequence) <= limit:
            return sequence
        return f"{sequence[:limit]}…{sequence[-min(limit, 40):]}"

    def _parse_pdb(self, path: Path, preview_limit: int) -> dict[str, Any]:
        """Parse a PDB text file in memory without running external software.

        The parser intentionally extracts the fields needed for evaluation,
        rather than attempting to repair or alter a structure. Only the first
        coordinate model is used for atom, sequence, and geometry summaries;
        all MODEL records are counted.
        """
        selected_model: int | None = None
        current_model = 1
        model_ids: set[int] = set()
        atom_count = 0
        hetero_atom_count = 0
        malformed_coordinate_records = 0
        chains: dict[str, list[dict[str, Any]]] = {}
        residue_lookup: dict[tuple[str, int, str, str], dict[str, Any]] = {}

        with path.open("r", encoding="utf-8", errors="replace") as handle:
            for line in handle:
                record_type = line[:6].strip().upper()
                if record_type == "MODEL":
                    try:
                        current_model = int(line[10:14].strip())
                    except ValueError:
                        current_model = len(model_ids) + 1
                    model_ids.add(current_model)
                    if selected_model is None:
                        selected_model = current_model
                    continue
                if record_type not in {"ATOM", "HETATM"}:
                    continue
                if selected_model is None:
                    selected_model = current_model
                model_ids.add(current_model)
                if current_model != selected_model:
                    continue
                atom_count += 1
                if record_type == "HETATM":
                    hetero_atom_count += 1
                    continue
                try:
                    residue_number = int(line[22:26].strip())
                    coordinate = (
                        float(line[30:38].strip()),
                        float(line[38:46].strip()),
                        float(line[46:54].strip()),
                    )
                    b_factor = float(line[60:66].strip()) if line[60:66].strip() else None
                except ValueError:
                    malformed_coordinate_records += 1
                    continue
                chain_id = line[21].strip() or "<blank>"
                residue_name = line[17:20].strip().upper() or "UNK"
                insertion_code = line[26].strip()
                atom_name = line[12:16].strip().upper()
                residue_key = (chain_id, residue_number, insertion_code, residue_name)
                residue = residue_lookup.get(residue_key)
                if residue is None:
                    residue = {
                        "chain_id": chain_id,
                        "name": residue_name,
                        "number": residue_number,
                        "insertion_code": insertion_code,
                        "code": THREE_TO_ONE.get(residue_name, "X"),
                        "atoms": {},
                    }
                    residue_lookup[residue_key] = residue
                    chains.setdefault(chain_id, []).append(residue)
                # First alternate location is deterministic and adequate for a
                # structure summary; no atom coordinates are written anywhere.
                residue["atoms"].setdefault(atom_name, {"coordinate": coordinate, "b_factor": b_factor})

        if selected_model is None:
            selected_model = 1
        chain_summaries: list[dict[str, Any]] = []
        residue_preview: list[dict[str, Any]] = []
        total_polymer_residues = 0
        total_ca_atoms = 0
        total_complete_backbones = 0
        for chain_id, residues in chains.items():
            sequence = "".join(residue["code"] for residue in residues)
            ca_b_factors = [
                atom["b_factor"]
                for residue in residues
                if (atom := residue["atoms"].get("CA")) is not None and atom["b_factor"] is not None
            ]
            ca_count = sum("CA" in residue["atoms"] for residue in residues)
            complete_backbone_count = sum(
                all(atom_name in residue["atoms"] for atom_name in ("N", "CA", "C"))
                for residue in residues
            )
            total_polymer_residues += len(residues)
            total_ca_atoms += ca_count
            total_complete_backbones += complete_backbone_count
            chain_summaries.append(
                {
                    "chain_id": chain_id,
                    "residue_count": len(residues),
                    "sequence_length": len(sequence),
                    "sequence_sha256": hashlib.sha256(sequence.encode("ascii")).hexdigest(),
                    "sequence_preview": self._sequence_preview(sequence, max(20, min(preview_limit, 120))),
                    "ca_atom_count": ca_count,
                    "complete_backbone_residue_count": complete_backbone_count,
                    "ca_b_factor_summary": self._summary(ca_b_factors),
                }
            )
            for residue in residues:
                if len(residue_preview) >= preview_limit:
                    break
                ca = residue["atoms"].get("CA")
                residue_preview.append(
                    {
                        "chain_id": chain_id,
                        "residue_name": residue["name"],
                        "residue_number": residue["number"],
                        "insertion_code": residue["insertion_code"],
                        "has_ca": ca is not None,
                        "has_complete_backbone": all(
                            atom_name in residue["atoms"] for atom_name in ("N", "CA", "C")
                        ),
                        "ca_b_factor": ca["b_factor"] if ca is not None else None,
                    }
                )

        summary = {
            "format": "pdb",
            "model_count": len(model_ids) or 1,
            "representative_model": selected_model,
            "chain_count": len(chains),
            "polymer_residue_count": total_polymer_residues,
            "atom_count_in_representative_model": atom_count,
            "hetero_atom_count_in_representative_model": hetero_atom_count,
            "ca_atom_count": total_ca_atoms,
            "complete_backbone_residue_count": total_complete_backbones,
            "malformed_coordinate_records": malformed_coordinate_records,
            "chains": chain_summaries,
            "residue_preview": residue_preview,
            "b_factor_note": (
                "CA B-factors are reported as stored. In AlphaFold-style predicted PDB files "
                "they commonly encode pLDDT, but this must not be assumed for every PDB."
            ),
        }
        return {"summary": summary, "chains": chains}

    def inspect_pdb(self, raw_path: str, residue_limit: int) -> dict[str, Any]:
        """Inspect one allowed PDB structure with an in-memory, read-only parser."""
        if residue_limit < 1:
            raise PolicyError("residue_limit must be positive")
        path, _ = self._resolve_allowed(raw_path)
        if path.suffix.lower() != PDB_EXTENSION:
            raise PolicyError("inspect_pdb only supports .pdb structure files")
        parsed = self._parse_pdb(path, min(residue_limit, self.limits.max_records))
        citation = self.cite(
            path,
            locator="PDB structure summary and residue preview",
            extractor="inspect_pdb",
            status="complete",
        )
        return {"path": str(path), **parsed["summary"], "evidence_id": citation.evidence_id}

    @staticmethod
    def _matching_blocks(reference_sequence: str, agent_sequence: str) -> list[tuple[int, int, int]]:
        return [
            (block.a, block.b, block.size)
            for block in SequenceMatcher(None, reference_sequence, agent_sequence, autojunk=False).get_matching_blocks()
            if block.size
        ]

    @staticmethod
    def _ca_rmsd_after_superposition(
        reference_points: list[tuple[float, float, float]],
        agent_points: list[tuple[float, float, float]],
    ) -> float | None:
        """Return Kabsch/Horn superposed CA-RMSD using only standard-library math."""
        if len(reference_points) != len(agent_points) or len(reference_points) < 3:
            return None
        count = len(reference_points)
        reference_center = tuple(sum(point[index] for point in reference_points) / count for index in range(3))
        agent_center = tuple(sum(point[index] for point in agent_points) / count for index in range(3))
        covariance = [[0.0] * 3 for _ in range(3)]
        for reference, agent in zip(reference_points, agent_points, strict=True):
            fixed = [reference[index] - reference_center[index] for index in range(3)]
            moving = [agent[index] - agent_center[index] for index in range(3)]
            for row in range(3):
                for column in range(3):
                    covariance[row][column] += moving[row] * fixed[column]
        sxx, sxy, sxz = covariance[0]
        syx, syy, syz = covariance[1]
        szx, szy, szz = covariance[2]
        horn = [
            [sxx + syy + szz, syz - szy, szx - sxz, sxy - syx],
            [syz - szy, sxx - syy - szz, sxy + syx, szx + sxz],
            [szx - sxz, sxy + syx, -sxx + syy - szz, syz + szy],
            [sxy - syx, szx + sxz, syz + szy, -sxx - syy + szz],
        ]
        # Shift the symmetric matrix before power iteration. This preserves its
        # eigenvectors while preventing equal-magnitude positive/negative
        # eigenvalues from making the iteration oscillate.
        shift = max(sum(abs(value) for value in row) for row in horn) + 1.0
        for index in range(4):
            horn[index][index] += shift
        quaternion = [1.0, 0.0, 0.0, 0.0]
        for _ in range(100):
            updated = [sum(horn[row][column] * quaternion[column] for column in range(4)) for row in range(4)]
            norm = math.sqrt(sum(value * value for value in updated))
            if norm == 0:
                return None
            updated = [value / norm for value in updated]
            delta = max(abs(updated[index] - quaternion[index]) for index in range(4))
            sign_flipped_delta = max(abs(updated[index] + quaternion[index]) for index in range(4))
            if min(delta, sign_flipped_delta) < 1e-12:
                quaternion = updated
                break
            quaternion = updated
        w, x, y, z = quaternion
        rotation = (
            (1 - 2 * (y * y + z * z), 2 * (x * y - w * z), 2 * (x * z + w * y)),
            (2 * (x * y + w * z), 1 - 2 * (x * x + z * z), 2 * (y * z - w * x)),
            (2 * (x * z - w * y), 2 * (y * z + w * x), 1 - 2 * (x * x + y * y)),
        )
        squared_error = 0.0
        for reference, agent in zip(reference_points, agent_points, strict=True):
            moving = [agent[index] - agent_center[index] for index in range(3)]
            rotated = [sum(rotation[row][column] * moving[column] for column in range(3)) for row in range(3)]
            for index in range(3):
                squared_error += (rotated[index] - (reference[index] - reference_center[index])) ** 2
        return round(math.sqrt(squared_error / count), 4)

    def compare_pdb_structures(self, raw_reference_path: str, raw_agent_path: str) -> dict[str, Any]:
        """Compare two allowed PDBs by chain sequence and superposed CA coordinates."""
        reference_path, _ = self._resolve_allowed(raw_reference_path)
        agent_path, _ = self._resolve_allowed(raw_agent_path)
        if reference_path.suffix.lower() != PDB_EXTENSION or agent_path.suffix.lower() != PDB_EXTENSION:
            raise PolicyError("compare_pdb_structures requires two .pdb files")
        reference = self._parse_pdb(reference_path, min(self.limits.max_records, 50))
        agent = self._parse_pdb(agent_path, min(self.limits.max_records, 50))

        available_agent_chains = dict(agent["chains"])
        mappings: list[dict[str, Any]] = []
        all_reference_points: list[tuple[float, float, float]] = []
        all_agent_points: list[tuple[float, float, float]] = []
        for reference_chain_id, reference_residues in reference["chains"].items():
            reference_sequence = "".join(item["code"] for item in reference_residues)
            candidates = []
            for agent_chain_id, agent_residues in available_agent_chains.items():
                agent_sequence = "".join(item["code"] for item in agent_residues)
                blocks = self._matching_blocks(reference_sequence, agent_sequence)
                matching_residues = sum(block[2] for block in blocks)
                similarity = matching_residues / max(len(reference_sequence), len(agent_sequence), 1)
                candidates.append((agent_chain_id, agent_residues, blocks, matching_residues, similarity))
            if not candidates:
                mappings.append(
                    {
                        "reference_chain_id": reference_chain_id,
                        "agent_chain_id": None,
                        "reference_sequence_length": len(reference_sequence),
                        "agent_sequence_length": 0,
                        "matching_residue_count": 0,
                        "sequence_identity_fraction": 0.0,
                        "ca_pair_count": 0,
                        "ca_rmsd_after_superposition": None,
                    }
                )
                continue
            # Keep matching chain IDs when possible; otherwise use the highest
            # exact-sequence correspondence among the still-unmatched chains.
            same_id_candidates = [candidate for candidate in candidates if candidate[0] == reference_chain_id]
            chosen = max(same_id_candidates or candidates, key=lambda item: (item[4], item[3], item[0]))
            agent_chain_id, agent_residues, blocks, matching_residues, similarity = chosen
            del available_agent_chains[agent_chain_id]
            reference_points: list[tuple[float, float, float]] = []
            agent_points: list[tuple[float, float, float]] = []
            for reference_start, agent_start, block_size in blocks:
                for offset in range(block_size):
                    reference_ca = reference_residues[reference_start + offset]["atoms"].get("CA")
                    agent_ca = agent_residues[agent_start + offset]["atoms"].get("CA")
                    if reference_ca is not None and agent_ca is not None:
                        reference_points.append(reference_ca["coordinate"])
                        agent_points.append(agent_ca["coordinate"])
            all_reference_points.extend(reference_points)
            all_agent_points.extend(agent_points)
            mappings.append(
                {
                    "reference_chain_id": reference_chain_id,
                    "agent_chain_id": agent_chain_id,
                    "reference_sequence_length": len(reference_sequence),
                    "agent_sequence_length": len(agent_residues),
                    "matching_residue_count": matching_residues,
                    "sequence_identity_fraction": round(similarity, 4),
                    "ca_pair_count": len(reference_points),
                    "ca_rmsd_after_superposition": self._ca_rmsd_after_superposition(reference_points, agent_points),
                }
            )

        reference_citation = self.cite(
            reference_path,
            locator=f"PDB comparison with {self._relative_path(agent_path)}",
            extractor="compare_pdb_structures",
            status="complete",
        )
        agent_citation = self.cite(
            agent_path,
            locator=f"PDB comparison with {self._relative_path(reference_path)}",
            extractor="compare_pdb_structures",
            status="complete",
        )
        return {
            "reference_path": str(reference_path),
            "agent_path": str(agent_path),
            "reference_structure": reference["summary"],
            "agent_structure": agent["summary"],
            "chain_mappings": mappings,
            "unmatched_agent_chain_ids": sorted(available_agent_chains),
            "overall_ca_pair_count": len(all_reference_points),
            "overall_ca_rmsd_after_superposition": self._ca_rmsd_after_superposition(
                all_reference_points,
                all_agent_points,
            ),
            "rmsd_note": "RMSD is after rigid-body superposition of sequence-matched CA atoms; it is descriptive evidence, not a fixed pass/fail threshold.",
            "reference_evidence_id": reference_citation.evidence_id,
            "agent_evidence_id": agent_citation.evidence_id,
            "evidence_ids": [reference_citation.evidence_id, agent_citation.evidence_id],
        }

    def _inspect_sequence_file(self, path: Path, record_limit: int) -> dict[str, Any]:
        try:
            from Bio import SeqIO
        except ImportError as exc:
            raise PolicyError("FASTA/FASTQ inspection requires biopython; install requirements.txt") from exc
        format_name = "fastq" if path.suffix.lower() in {".fq", ".fastq"} else "fasta"
        records = []
        for count, record in enumerate(SeqIO.parse(path, format_name), start=1):
            if count > record_limit:
                break
            item: dict[str, Any] = {"id": record.id, "description": record.description, "length": len(record.seq), "sequence_preview": str(record.seq)[:120]}
            if format_name == "fastq":
                qualities = record.letter_annotations.get("phred_quality", [])
                item["quality_length"] = len(qualities)
                item["quality_preview"] = qualities[:20]
            records.append(item)
        citation = self.cite(path, locator=f"first {len(records)} records", extractor="inspect_bio", status="partial", notes=["Record inspection is sampled."])
        return {"path": str(path), "format": format_name, "sample_records": records, "truncated": True, "evidence_id": citation.evidence_id}

    def _inspect_variant_file(self, path: Path, record_limit: int) -> dict[str, Any]:
        try:
            import pysam
        except ImportError as exc:
            raise PolicyError("VCF/BCF inspection requires pysam; install requirements.txt") from exc
        variants = pysam.VariantFile(path)
        records = []
        for count, record in enumerate(variants, start=1):
            if count > record_limit:
                break
            records.append({"contig": record.contig, "pos": record.pos, "id": record.id, "ref": record.ref, "alts": list(record.alts or []), "qual": record.qual, "filter": list(record.filter.keys()), "info": dict(record.info)})
        header = {"samples": list(variants.header.samples), "contigs": list(variants.header.contigs)[:100], "formats": list(variants.header.formats)[:100], "infos": list(variants.header.info)[:100]}
        variants.close()
        citation = self.cite(path, locator=f"header and first {len(records)} records", extractor="inspect_bio", status="partial", notes=["Variant inspection is sampled."])
        return {"path": str(path), "format": "vcf_or_bcf", "header": header, "sample_records": records, "truncated": True, "evidence_id": citation.evidence_id}

    def _inspect_alignment_file(self, path: Path, record_limit: int) -> dict[str, Any]:
        try:
            import pysam
        except ImportError as exc:
            raise PolicyError("BAM/CRAM inspection requires pysam; install requirements.txt") from exc
        mode = "rb" if path.suffix.lower() == ".bam" else "rc"
        with pysam.AlignmentFile(path, mode) as alignments:
            records = []
            for count, record in enumerate(alignments.fetch(until_eof=True), start=1):
                if count > record_limit:
                    break
                records.append({"query_name": record.query_name, "flag": record.flag, "reference_name": record.reference_name, "reference_start": record.reference_start, "mapping_quality": record.mapping_quality, "cigar": record.cigarstring, "query_length": record.query_length})
            header = alignments.header.to_dict()
        citation = self.cite(path, locator=f"header and first {len(records)} alignment records", extractor="inspect_bio", status="partial", notes=["Alignment inspection is sampled; no index or command execution is used."])
        return {"path": str(path), "format": "bam_or_cram", "header": header, "sample_records": records, "truncated": True, "evidence_id": citation.evidence_id}

    @property
    def records(self) -> list[EvidenceRecord]:
        with self._lock:
            return list(self._records)

    @property
    def valid_evidence_ids(self) -> set[str]:
        with self._lock:
            return {record.evidence_id for record in self._records}

    def was_content_inspected(self, path: Path) -> bool:
        """True when a file was inspected with a content-reading tool."""
        resolved = path.resolve()
        with self._lock:
            return any(record.path == self._relative_path(resolved) for record in self._records)

    def coverage(self) -> dict[str, Any]:
        with self._lock:
            all_paths = set(self._files)
            inspected_paths = {record.path for record in self._records}
            inspected = {path for path in all_paths if self._relative_path(path) in inspected_paths}
            return {
                "files_discovered": len(all_paths),
                "files_inspected": len(inspected),
                "files_not_inspected": [self._relative_path(path) for path in sorted(all_paths - inspected)],
                "partial_evidence_paths": [self._relative_path(path) for path in sorted(self._partial_paths)],
            }
